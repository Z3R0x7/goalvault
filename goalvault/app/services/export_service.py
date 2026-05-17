import io
from datetime import date
from flask import current_app
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from app.models.goal import Goal
from app.models.achievement import Achievement
from app.models.user import User
from app.services.analytics_service import AnalyticsService


class ExportService:
    @staticmethod
    def _score_fill(score):
        if score is None or score == "":
            return PatternFill("solid", fgColor="484F58")
        s = float(score)
        if s >= 80:
            return PatternFill("solid", fgColor="238636")
        if s >= 50:
            return PatternFill("solid", fgColor="9E6A03")
        return PatternFill("solid", fgColor="DA3633")

    @staticmethod
    def generate_achievement_report(cycle_year=None):
        year = cycle_year or current_app.config.get("CYCLE_YEAR", 2025)
        wb = Workbook()
        summary = wb.active
        summary.title = "Summary"
        summary["A1"] = "GoalVault Achievement Report"
        summary["A1"].font = Font(bold=True, size=16, color="00C8A0")
        summary["A2"] = f"Cycle Year: {year}  |  Generated: {date.today()}"
        analytics = AnalyticsService.get_org_summary(year)
        summary["A4"] = "Total Employees"
        summary["B4"] = analytics["totals"]["employees"]
        summary["A5"] = "Submission Rate"
        summary["B5"] = f"{analytics['totals']['submitted_pct']}%"
        summary["A6"] = "Approval Rate"
        summary["B6"] = f"{analytics['totals']['approved_pct']}%"
        summary["A7"] = "Check-in Rate"
        summary["B7"] = f"{analytics['totals']['checkin_pct']}%"

        detail = wb.create_sheet("Employee Detail")
        headers = [
            "Employee", "Department", "Goal", "UoM", "Target", "Weightage",
            "Q1 Actual", "Q1 Progress", "Q2 Actual", "Q2 Progress",
            "Q3 Actual", "Q3 Progress", "Q4 Actual", "Q4 Progress", "Status",
        ]
        header_fill = PatternFill("solid", fgColor="161B22")
        header_font = Font(bold=True, color="E6EDF3")
        for col, h in enumerate(headers, 1):
            cell = detail.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font

        row_idx = 2
        employees = User.query.filter_by(role="employee", is_active=True).all()
        for emp in employees:
            goals = Goal.query.filter_by(
                employee_id=emp.id, cycle_year=year
            ).filter(Goal.status.in_(["approved", "locked"])).all()
            for goal in goals:
                vals = [emp.name, emp.department or "", goal.title, goal.uom_type, goal.target, goal.weightage]
                progress_indices = []
                for q in ["Q1", "Q2", "Q3", "Q4"]:
                    ach = Achievement.query.filter_by(goal_id=goal.id, quarter=q).first()
                    vals.append(ach.actual_value if ach else "")
                    vals.append(ach.progress_score if ach else "")
                    progress_indices.append((len(vals), ach.progress_score if ach else None))
                vals.append(goal.status)
                for col, val in enumerate(vals, 1):
                    detail.cell(row=row_idx, column=col, value=val)
                for col_i, score in progress_indices:
                    cell = detail.cell(row=row_idx, column=col_i)
                    if score is not None:
                        cell.fill = ExportService._score_fill(score)
                row_idx += 1

        if row_idx > 2:
            detail.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_idx - 1}"
        detail.freeze_panes = "A2"

        dept_sheet = wb.create_sheet("Department Summary")
        dept_sheet.append(["Department", "Approved %", "Q1", "Q2", "Q3", "Q4"])
        for row in analytics["heatmap"]:
            dept_sheet.append([
                row["department"],
                row["approved_pct"],
                row["quarters"].get("Q1") or "—",
                row["quarters"].get("Q2") or "—",
                row["quarters"].get("Q3") or "—",
                row["quarters"].get("Q4") or "—",
            ])

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output, f"goalvault_achievement_report_{year}.xlsx"
